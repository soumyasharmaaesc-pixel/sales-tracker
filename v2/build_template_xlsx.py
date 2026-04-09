#!/usr/bin/env python3
"""Generate the empty benchmark template xlsx that users can download,
fill in, and re-upload to the dashboard."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "/sessions/lucid-vibrant-ramanujan/sales-dashboard/vercel-deploy"
OUT_PATH = os.path.join(OUT_DIR, "sales-tracker-benchmark-template.xlsx")

# ─── Styling helpers ──────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1E1B4B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
SECTION_FILL = PatternFill("solid", start_color="EEF2FF")
SECTION_FONT = Font(bold=True, color="1E1B4B", size=11, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", size=9, italic=True, color="6B7280")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ─── Sheet 1: Instructions ────────────────────────────────────────────
ws0 = wb.active
ws0.title = "Instructions"
ws0.column_dimensions["A"].width = 110

ws0["A1"] = "TheHireHub.AI · Sales Tracker Benchmark Template"
ws0["A1"].font = Font(bold=True, size=16, color="1E1B4B", name="Arial")

instructions = [
    "",
    "HOW TO USE THIS TEMPLATE",
    "",
    "1. Fill in your weekly targets on the 'Weekly Targets' sheet — one row per week.",
    "   You can add or remove rows as needed, but keep the column headers intact.",
    "",
    "2. Optionally, fill in your daily targets on the 'Daily Targets' sheet.",
    "   Only use this if you want day-level granularity; weekly rollup is used by default.",
    "",
    "3. Save the file as .xlsx (Excel) or keep as .xls / .csv — all three are supported.",
    "",
    "4. In the dashboard, go to Manager View → GTM Benchmarks tab → click 'Upload Benchmark'",
    "   and select this file. The rep's weekly submissions will be scored against these targets.",
    "",
    "5. Uploading a new benchmark does NOT remove old ones. Every upload is kept in the",
    "   dashboard's history reservoir so you can see how targets changed over time.",
    "",
    "WHICH FIELDS ARE REQUIRED",
    "",
    "All numeric columns are optional — leave blank if you don't want to set a target.",
    "At minimum, fill in the fields you want to track (most teams track Calls, Emails,",
    "Demos, Proposals, and Deals Closed).",
    "",
    "COLUMN DEFINITIONS",
    "",
    "Week                   — Sequential week number (1, 2, 3, ...)",
    "Week Label             — Human-readable label, e.g. 'Week 1 (Apr 10–16)'",
    "Calls Target           — Target number of discovery / outbound calls for the week",
    "Emails Target          — Target number of cold / follow-up emails sent",
    "LinkedIn Touches       — Target number of LinkedIn messages + connection requests",
    "Demos Booked           — Target number of demos scheduled on the calendar",
    "Demos Done             — Target number of demos actually delivered",
    "Proposals Sent         — Target number of proposals / SOWs sent to prospects",
    "Deals Closed           — Target number of deals closed-won for the week",
    "MRR ($) Target         — Target new monthly recurring revenue for the week ($)",
    "Pipeline Value ($)     — Target total open pipeline value by end of week ($)",
    "Notes                  — Free-text notes, context, stretch/commit targets, etc.",
]
for i, line in enumerate(instructions, start=2):
    cell = ws0.cell(row=i, column=1, value=line)
    if line.isupper() and len(line) > 3:
        cell.font = Font(bold=True, size=11, color="4F46E5", name="Arial")
    elif line and not line.startswith(" "):
        cell.font = BODY_FONT
    else:
        cell.font = BODY_FONT
    cell.alignment = LEFT

# ─── Sheet 2: Weekly Targets ──────────────────────────────────────────
ws1 = wb.create_sheet("Weekly Targets")

weekly_headers = [
    "Week", "Week Label",
    "Calls Target", "Emails Target", "LinkedIn Touches",
    "Demos Booked", "Demos Done", "Proposals Sent", "Deals Closed",
    "MRR ($) Target", "Pipeline Value ($)", "Notes",
]
widths = [7, 24, 13, 14, 18, 14, 13, 15, 14, 16, 20, 32]

for col_idx, (header, width) in enumerate(zip(weekly_headers, widths), start=1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# Pre-populate Weeks 1–12 with empty target cells so the user can just fill in
from datetime import date, timedelta
start = date(2026, 4, 10)  # Day 1

for wk in range(1, 13):
    row = wk + 1
    week_start = start + timedelta(days=(wk - 1) * 7)
    week_end = week_start + timedelta(days=6)
    label = f"Week {wk} ({week_start.strftime('%b %d')}–{week_end.strftime('%d')})"

    ws1.cell(row=row, column=1, value=wk).alignment = CENTER
    ws1.cell(row=row, column=2, value=label)
    # Numeric columns left blank for the user to fill in
    for col in range(3, 13):
        c = ws1.cell(row=row, column=col, value=None)
        c.alignment = CENTER
        if col in (10, 11):  # $ columns
            c.number_format = '"$"#,##0;[Red]-"$"#,##0;"-"'
        else:
            c.number_format = '#,##0;[Red]-#,##0;"-"'

    for col in range(1, 13):
        ws1.cell(row=row, column=col).border = BORDER
        if ws1.cell(row=row, column=col).font is None:
            ws1.cell(row=row, column=col).font = BODY_FONT

# Totals row
total_row = 14
ws1.cell(row=total_row, column=1, value="").fill = SECTION_FILL
ws1.cell(row=total_row, column=2, value="TOTAL (12 weeks)").font = SECTION_FONT
ws1.cell(row=total_row, column=2).fill = SECTION_FILL
for col in range(3, 12):
    col_letter = get_column_letter(col)
    formula = f"=SUM({col_letter}2:{col_letter}13)"
    cell = ws1.cell(row=total_row, column=col, value=formula)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = CENTER
    if col in (10, 11):
        cell.number_format = '"$"#,##0;[Red]-"$"#,##0;"-"'
    else:
        cell.number_format = '#,##0;[Red]-#,##0;"-"'
    cell.border = BORDER
ws1.cell(row=total_row, column=1).border = BORDER
ws1.cell(row=total_row, column=2).border = BORDER
ws1.cell(row=total_row, column=12, value="").fill = SECTION_FILL
ws1.cell(row=total_row, column=12).border = BORDER

ws1.freeze_panes = "A2"

# ─── Sheet 3: Daily Targets ───────────────────────────────────────────
ws2 = wb.create_sheet("Daily Targets")
daily_headers = [
    "Day", "Date", "Calls", "Emails", "LinkedIn Touches",
    "Demos Booked", "Demos Done", "Notes",
]
d_widths = [6, 13, 10, 10, 17, 14, 13, 32]
for col_idx, (header, width) in enumerate(zip(daily_headers, d_widths), start=1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

# 20 working days of rows (4 weeks × 5 days)
day_counter = 0
d = start
while day_counter < 20:
    if d.weekday() < 5:  # Mon-Fri
        day_counter += 1
        row = day_counter + 1
        ws2.cell(row=row, column=1, value=day_counter).alignment = CENTER
        date_cell = ws2.cell(row=row, column=2, value=d)
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.alignment = CENTER
        for col in range(3, 8):
            c = ws2.cell(row=row, column=col, value=None)
            c.alignment = CENTER
            c.number_format = '#,##0;[Red]-#,##0;"-"'
        for col in range(1, 9):
            ws2.cell(row=row, column=col).border = BORDER
            if ws2.cell(row=row, column=col).font is None:
                ws2.cell(row=row, column=col).font = BODY_FONT
    d += timedelta(days=1)

# Daily totals row
d_total_row = 22
ws2.cell(row=d_total_row, column=1, value="").fill = SECTION_FILL
ws2.cell(row=d_total_row, column=2, value="TOTAL (20 days)").font = SECTION_FONT
ws2.cell(row=d_total_row, column=2).fill = SECTION_FILL
for col in range(3, 8):
    col_letter = get_column_letter(col)
    formula = f"=SUM({col_letter}2:{col_letter}21)"
    cell = ws2.cell(row=d_total_row, column=col, value=formula)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = CENTER
    cell.number_format = '#,##0;[Red]-#,##0;"-"'
    cell.border = BORDER
ws2.cell(row=d_total_row, column=1).border = BORDER
ws2.cell(row=d_total_row, column=2).border = BORDER
ws2.cell(row=d_total_row, column=8, value="").fill = SECTION_FILL
ws2.cell(row=d_total_row, column=8).border = BORDER

ws2.freeze_panes = "A2"

# Tip row at the bottom
tip_row = d_total_row + 2
ws2.cell(row=tip_row, column=1, value="TIP:").font = Font(bold=True, color="4F46E5", name="Arial", size=10)
ws2.cell(
    row=tip_row, column=2,
    value="Leave this sheet blank if you only want weekly targets. The dashboard uses daily targets only if at least one row has values."
).font = NOTE_FONT
ws2.merge_cells(start_row=tip_row, start_column=2, end_row=tip_row, end_column=8)

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}: {os.path.getsize(OUT_PATH)} bytes")
